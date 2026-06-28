"""XLSX-экспортёр результатов BDU-сканирования и SBOM-компонентов."""
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


_SEV_COLORS = {
    "critical": "C00000",
    "high":     "FF0000",
    "medium":   "FF9900",
    "low":      "FFFF00",
    "unknown":  "D9D9D9",
}

_STATUS_COLORS = {
    "open":           "FF9900",
    "confirmed":      "FF0000",
    "resolved":       "70AD47",
    "risk_accepted":  "4472C4",
    "false_positive": "A9A9A9",
}


def _header_style(ws, row, cols):
    fill = PatternFill("solid", fgColor="1F1F2E")
    font = Font(bold=True, color="FFFFFF", size=10)
    for col, title in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def export_bdu_scan(scan_result: dict, sbom_stats: dict, out_path: Path) -> bool:
    """
    Экспортирует результаты BDU-сканирования в XLSX.
    Возвращает True при успехе.
    """
    if not OPENPYXL_OK:
        return False

    wb = openpyxl.Workbook()

    # ── Лист 1: Сводка ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Сводка"
    ws.row_dimensions[1].height = 30
    stats = scan_result.get("stats", {})
    summary_data = [
        ("Всего компонентов проверено", scan_result.get("total_components_scanned", 0)),
        ("Компонентов с уязвимостями", stats.get("affected_components", 0)),
        ("Всего уязвимостей найдено",  stats.get("total_vulns", 0)),
        ("Критических (Critical)",      stats.get("critical", 0)),
        ("Высоких (High)",              stats.get("high", 0)),
        ("Средних (Medium)",            stats.get("medium", 0)),
        ("Низких (Low)",                stats.get("low", 0)),
        ("С известным эксплойтом",     stats.get("with_exploit", 0)),
        ("Всего компонентов в SBOM",   sbom_stats.get("total", 0)),
        ("Компонентов с PURL",         sbom_stats.get("with_purl", 0)),
        ("Компонентов с версией",      sbom_stats.get("with_version", 0)),
        ("Компонентов с лицензией",    sbom_stats.get("with_license", 0)),
    ]
    ws["A1"] = "Параметр"
    ws["B1"] = "Значение"
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["B1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F1F2E")
    ws["B1"].fill = PatternFill("solid", fgColor="1F1F2E")
    for i, (label, value) in enumerate(summary_data, 2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        if "Critical" in label or "критич" in label.lower():
            ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor="FFCCCC")
        elif "High" in label or "высок" in label.lower():
            ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor="FFE0CC")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    # ── Лист 2: Уязвимости ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Уязвимости")
    cols = ["Компонент", "Версия", "BDU ID", "CVE ID", "Критичность",
            "CVSS", "Эксплойт", "Статус триажа", "Приоритет", "Решение",
            "Дата публикации", "PURL", "Тип сопоставления"]
    widths = [22, 12, 18, 16, 12, 8, 10, 14, 10, 30, 15, 30, 16]
    _header_style(ws2, 1, cols)
    _col_widths(ws2, widths)
    ws2.freeze_panes = "A2"
    ws2.row_dimensions[1].height = 25

    row = 2
    for comp_entry in scan_result.get("components", []):
        comp = comp_entry.get("component", {})
        for v in comp_entry.get("vulnerabilities", []):
            sev = v.get("severity", "unknown")
            status = v.get("status", "open")
            has_exp = "Да" if v.get("has_exploit") else "Нет"
            pub_exp = " (публичный)" if v.get("public_exploit") else ""

            cells = [
                comp.get("name", ""),
                comp.get("version", ""),
                v.get("bdu_id", ""),
                ", ".join(v.get("cve_ids", [])),
                sev.upper(),
                v.get("cvss_score", ""),
                has_exp + pub_exp,
                status,
                v.get("priority_score", ""),
                (v.get("fix_status") or "") + " " + (v.get("solution") or "")[:100],
                v.get("pub_date", ""),
                comp.get("purl", ""),
                v.get("match_type", ""),
            ]
            for col, val in enumerate(cells, 1):
                c = ws2.cell(row=row, column=col, value=val)
                c.alignment = Alignment(wrap_text=False, vertical="top")

            # Color by severity
            sev_color = _SEV_COLORS.get(sev, "D9D9D9")
            for col in range(1, 14):
                ws2.cell(row=row, column=5).fill = PatternFill("solid", fgColor=sev_color)

            # Color status
            status_color = _STATUS_COLORS.get(status)
            if status_color:
                ws2.cell(row=row, column=8).fill = PatternFill("solid", fgColor=status_color)
                ws2.cell(row=row, column=8).font = Font(color="FFFFFF", bold=True)

            row += 1

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    # ── Лист 3: Компоненты SBOM ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Компоненты SBOM")
    comp_cols = ["Название", "Версия", "Тип", "PURL", "Лицензия"]
    comp_widths = [30, 16, 14, 50, 20]
    _header_style(ws3, 1, comp_cols)
    _col_widths(ws3, comp_widths)
    ws3.freeze_panes = "A2"

    for i, c in enumerate(sbom_stats.get("components", []), 2):
        ws3.cell(row=i, column=1, value=c.get("name", ""))
        ws3.cell(row=i, column=2, value=c.get("version", ""))
        ws3.cell(row=i, column=3, value=c.get("type", ""))
        ws3.cell(row=i, column=4, value=c.get("purl", ""))
        ws3.cell(row=i, column=5, value=c.get("licenses", ""))
    ws3.auto_filter.ref = f"A1:E1"

    wb.save(out_path)
    return True
